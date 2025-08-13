#!/usr/bin/env python3
# Excel -> ICS generator with all-day support and "free/busy" control.
#
# Expected headers in row 1:
# Unique ID | Course Code | Title | Category | Start Date | Start Time | End Date | End Time
# Timezone | Location | Description | Link | TRANSPARENT
#
# Usage:
#   python generate_ics.py --xlsx calendar.xlsx --out docs/calendar.ics
#
# Notes:
# - All-day: leave Start Time and End Time blank. DTEND (DATE) is exclusive,
#   so we add +1 day to the end date internally to make user-specified End Date inclusive.
# - Timed events: supply Start Time and End Time (24h HH:MM). Timezone defaults to Australia/Sydney.
# - TRANSPARENT column: TRUE/YES/TRANSPARENT => TRANSP:TRANSPARENT else OPAQUE.
# - If you later add a "RRULE" or "EXDATE" header, this script will include them automatically.
# - Includes a VTIMEZONE for Australia/Sydney (add more as needed).

import argparse
from openpyxl import load_workbook
from datetime import datetime, timedelta
import hashlib
import os, sys

# --- Config ---
DEFAULT_TZ = "Australia/Sydney"

AUS_TZ_VTIMEZONE = """BEGIN:VTIMEZONE
TZID:Australia/Sydney
BEGIN:STANDARD
DTSTART:19700405T030000
TZOFFSETFROM:+1100
TZOFFSETTO:+1000
TZNAME:AEST
RRULE:FREQ=YEARLY;BYMONTH=4;BYDAY=1SU
END:STANDARD
BEGIN:DAYLIGHT
DTSTART:19701004T020000
TZOFFSETFROM:+1000
TZOFFSETTO:+1100
TZNAME:AEDT
RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=1SU
END:DAYLIGHT
END:VTIMEZONE
"""

def fmt_local(dt: datetime) -> str:
    return dt.strftime("%Y%m%dT%H%M%S")

def fmt_date(d: datetime) -> str:
    return d.strftime("%Y%m%d")

def parse_date(s):
    if not s:
        return None
    return datetime.strptime(str(s).strip(), "%Y-%m-%d")

def parse_time(s):
    if not s:
        return None
    return datetime.strptime(str(s).strip(), "%H:%M")

def parse_datetime(date_str, time_str):
    d = parse_date(date_str)
    if d is None:
        return None
    if not time_str:
        # default to midnight if time omitted (only used for timed events fallback)
        return datetime(d.year, d.month, d.day, 0, 0, 0)
    t = parse_time(time_str)
    return datetime(d.year, d.month, d.day, t.hour, t.minute, 0)

def truthy(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in {"true", "yes", "y", "1", "transparent", "free"}

def make_uid(fields):
    h = hashlib.sha1("|".join(fields).encode("utf-8")).hexdigest()[:16]
    return f"{h}@youruni"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default="Events")  # if your sheet is named differently, set here
    ap.add_argument("--out", default="docs/calendar.ics")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"Sheet '{args.sheet}' not found", file=sys.stderr)
        sys.exit(1)
    ws = wb[args.sheet]

    # Header map (case-insensitive match)
    hdr_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [ (h or "").strip() for h in hdr_cells ]
    header_map = { h.lower(): i for i, h in enumerate(headers) }

    def col(name):
        return header_map.get(name.lower(), -1)

    col_UID = col("Unique ID")
    col_Course = col("Course Code")
    col_Title = col("Title")
    col_Cat = col("Category")
    col_SDate = col("Start Date")
    col_STime = col("Start Time")
    col_EDate = col("End Date")
    col_ETime = col("End Time")
    col_TZ = col("Timezone")
    col_Loc = col("Location")
    col_Desc = col("Description")
    col_URL = col("Link")
    col_TRANSP = col("TRANSPARENT")

    # Optional extra columns if you add them later
    col_RRULE = header_map.get("rrule", -1)
    col_EXDATE = header_map.get("exdate", -1)

    now_utc = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    lines = [
        "BEGIN:VCALENDAR",
        "PRODID:-//YourUni//Class Feeds 1.0//EN",
        "VERSION:2.0",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        AUS_TZ_VTIMEZONE.strip()
    ]

    for r in ws.iter_rows(min_row=2, values_only=True):
        # skip entirely empty rows
        if r is None or all(v in (None, "") for v in r):
            continue

        title = (r[col_Title] or "").strip() if col_Title >= 0 and r[col_Title] else ""
        if not title:
            continue

        uid = (r[col_UID] or "").strip() if col_UID >= 0 and r[col_UID] else ""
        course = (r[col_Course] or "").strip() if col_Course >= 0 and r[col_Course] else ""
        cat = (r[col_Cat] or "").strip() if col_Cat >= 0 and r[col_Cat] else ""
        sdate = (r[col_SDate] or "").strip() if col_SDate >= 0 and r[col_SDate] else ""
        stime = (r[col_STime] or "").strip() if col_STime >= 0 and r[col_STime] else ""
        edate = (r[col_EDate] or "").strip() if col_EDate >= 0 and r[col_EDate] else ""
        etime = (r[col_ETime] or "").strip() if col_ETime >= 0 and r[col_ETime] else ""
        tz = (r[col_TZ] or "").strip() if col_TZ >= 0 and r[col_TZ] else DEFAULT_TZ
        location = (r[col_Loc] or "").strip() if col_Loc >= 0 and r[col_Loc] else ""
        desc = (r[col_Desc] or "").strip() if col_Desc >= 0 and r[col_Desc] else ""
        url = (r[col_URL] or "").strip() if col_URL >= 0 and r[col_URL] else ""
        is_transparent = truthy(r[col_TRANSP]) if col_TRANSP >= 0 else False

        rrule = (r[col_RRULE] or "").strip() if col_RRULE >= 0 and r[col_RRULE] else ""
        exdate_raw = (r[col_EXDATE] or "").strip() if col_EXDATE >= 0 and r[col_EXDATE] else ""

        if not sdate:
            # must have at least a start date
            continue

        # Decide all-day vs timed
        is_all_day = (not stime and not etime)

        # Build UID if missing
        if not uid:
            base_fields = [course, title, sdate, edate or "", stime or "", etime or "", location]
            uid = make_uid(base_fields)

        summary = f"{course} — {title}" if course else title

        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        lines.append(f"DTSTAMP:{now_utc}")
        lines.append(f"SUMMARY:{summary}")
        if location:
            lines.append(f"LOCATION:{location}")
        if desc:
            # Preserve \n as literal newlines in ICS
            lines.append("DESCRIPTION:" + desc.replace("\\n", "\\n"))
        if url:
            lines.append(f"URL:{url}")

        # Categories for filtering/colouring
        cats = []
        if course: cats.append(course)
        if cat: cats.append(cat)
        if location: cats.append(location)
        if cats:
            lines.append(f"CATEGORIES:{','.join(cats)}")

        # Free/Busy transparency
        lines.append(f"TRANSP:{'TRANSPARENT' if is_transparent else 'OPAQUE'}")

        if is_all_day:
            # All-day: DATE values (no TZ). DTEND is exclusive, so add +1 day.
            start_d = parse_date(sdate)
            if not start_d:
                lines.append("END:VEVENT")
                continue

            if edate:
                end_d_inclusive = parse_date(edate)
                if not end_d_inclusive:
                    end
